"""
export_utils.py - Xuất kết quả phương án cọc ra nhiều định dạng tệp.

Module này gom ba kiểu xuất phục vụ trình bày / lưu trữ kết quả tối ưu:
    - Excel (.xlsx)  : bảng số liệu đầu vào, phương án kiến nghị, bảng kiểm
                       tra nội lực cọc theo từng tổ hợp và bảng tọa độ cọc.
    - PDF (_report)  : báo cáo tóm tắt 1 trang (thông số, phương án, hình
                       mặt bằng nhúng kèm).
    - PNG (_plan)    : ảnh mặt bằng bố trí cọc (lưu nguyên khổ figure).

Quy ước đơn vị: tải trọng nhập theo kN / kN.m; sức chịu tải và lực cọc quy
về Tấn (T) để so với [Po] / [Ct]. Lực cọc tính bằng công thức bệ cứng cho ra
đơn vị thô (theo N của tải, thường kN), nên trước khi so sánh phải nhân hệ số
hiệu chỉnh calib = config['pmax'] / (Pmax thô lớn nhất) để quy về cùng đơn vị
với config['pmax'] (đã ở Tấn).
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime

from core import rigid_cap


# ============================================================================
# Tiện ích chung
# ============================================================================
def calculate_pile_forces(coords, loads):
    """
    Lực từng cọc theo bệ cứng, trả về {chỉ_số_tổ_hợp: [P_cọc...]}.

    Tham số:
        coords : tọa độ cọc (n, 2).
        loads  : danh sách tổ hợp tải trọng.
    Dùng chung công thức với core.rigid_cap (tránh viết lặp). Trả về dict
    rỗng nếu không có cọc nào.
    """
    if len(coords) == 0:
        return {}
    P = rigid_cap.forces_all_loads(coords, loads)   # (len(loads), n)
    return {i: P[i].tolist() for i in range(len(loads))}


# ============================================================================
# XUẤT PNG
# ============================================================================
def export_png(plot_canvas_instance, coords, params, out_dir, prefix):
    """
    Lưu ảnh mặt bằng bố trí cọc ra tệp PNG.

    Tham số:
        plot_canvas_instance : canvas vẽ sẵn (PlotCanvas) để tái dùng figure.
        coords               : tọa độ cọc cần vẽ.
        params               : thông số bệ/cọc dùng khi vẽ.
        out_dir, prefix      : thư mục đích và tiền tố tên tệp.
    Tạo tệp '<prefix>_plan.png'. Trả về đường dẫn tệp đã ghi.
    """
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    out_path = os.path.join(out_dir, f"{prefix}_plan.png")

    # Tái dùng đúng canvas đang có nhưng lưu lại thành ảnh.
    # Lưu ý: plot_canvas_instance phải đã được vẽ trước đó!
    # KHÔNG dùng bbox_inches='tight': lưu nguyên khổ figure cố định để mọi ảnh
    # có cùng kích thước, tránh ảnh không đồng đều khi gộp vào PDF tổng hợp.
    plot_canvas_instance.draw_simulation(coords, params)
    plot_canvas_instance.fig.savefig(out_path, dpi=200)
    return out_path


# ============================================================================
# XUẤT EXCEL
# ============================================================================
def export_excel(config, loads, params, out_dir, prefix):
    """
    Xuất bảng kết quả phương án cọc ra tệp Excel (.xlsx).

    Tham số:
        config          : phương án kiến nghị (type, n, pmax, pmin, coords...).
        loads           : danh sách tổ hợp tải trọng.
        params          : thông số đầu vào (kích thước bệ, đường kính, [Po]...).
        out_dir, prefix : thư mục đích và tiền tố tên tệp.
    Tạo tệp '<prefix>_result.xlsx' gồm: thông số đầu vào, phương án kiến
    nghị, bảng kiểm tra nội lực cọc theo từng tổ hợp và bảng tọa độ cọc.
    Trả về đường dẫn tệp đã ghi.
    """
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    out_path = os.path.join(out_dir, f"{prefix}_result.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ket Qua"

    # Định dạng dùng chung (đậm / nền tiêu đề / canh giữa)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Tiêu đề báo cáo (gộp ô A1:G1)
    ws.merge_cells('A1:G1')
    ws['A1'] = f"BAO CAO TOI UU HOA - {prefix}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    # Khối thông số đầu vào
    ws['A3'] = "Thong so dau vao:"
    ws['A3'].font = bold_font
    ws.append(['Lx (m)', params.get('L_X', 0), 'Ly (m)', params.get('L_Y', 0)])
    ws.append(['d (m)', params.get('D_PILE', 0), 'SAFE_D (m)', params.get('SAFE_D', 0)])
    ws.append(['Po cho phep (T)', params.get('P_LIMIT', 0), 'Ct nho (T)', params.get('P_TENSION', 0)])

    # Khối phương án kiến nghị
    ws.append([])
    ws['A8'] = "Phuong an kien nghi:"
    ws['A8'].font = bold_font
    ws.append(['Kieu', config.get('type', 'Unknown')])
    ws.append(['So luong coc', config.get('n', 0)])
    ws.append(['Pmax (T)', round(config.get('pmax', 0), 2)])
    ws.append(['Pmin (T)', round(config.get('pmin', 0), 2)])
    ws.append(['Trang thai', 'DAT' if config.get('ok') else 'KHONG DAT'])
    ws.append(['Ly do', config.get('msg', '')])

    # Bảng kiểm tra nội lực cọc - tiêu đề khối (gộp ô, in đậm, nền)
    ws.append([])
    row = ws.max_row + 1
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "BANG KIEM TRA NOI LUC COC"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align

    # Hàng tiêu đề cột của bảng nội lực
    headers = ["Load Case", "N (T)", "Mx (T.m)", "My (T.m)", "Pmax (T)", "Pmin (T)", "Trang thai"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align

    forces_by_load = calculate_pile_forces(config['coords'], loads)

    # Lực cọc tính theo công thức bệ cứng cho ra đơn vị thô (theo N của tải, thường kN),
    # trong khi config['pmax'] đã được hiệu chỉnh về cùng đơn vị với [Po] (T).
    # => Quy về cùng đơn vị bằng hệ số: calib = config['pmax'] / (Pmax thô lớn nhất).
    raw_pmax = max((max(v) for v in forces_by_load.values() if v), default=0.0)
    cfg_pmax = config.get('pmax', 0) or 0
    calib = (cfg_pmax / raw_pmax) if (raw_pmax > 0 and cfg_pmax > 0) else 1.0

    # Ghi từng tổ hợp: quy lực cọc về Tấn rồi so với [Po] / [Ct]
    P_LIMIT = params.get('P_LIMIT', 900)
    P_TENSION = params.get('P_TENSION', 0)
    for i, load in enumerate(loads):
        piles_p = forces_by_load.get(i, [])
        pmax = (max(piles_p) if piles_p else 0) * calib
        pmin = (min(piles_p) if piles_p else 0) * calib
        status = "DAT" if (pmax <= P_LIMIT and pmin >= -P_TENSION) else "FAIL"
        ws.append([i+1, load.get('N',0), load.get('Mx',0), load.get('My',0), round(pmax,2), round(pmin,2), status])

    # Bảng tọa độ cọc - tiêu đề khối (gộp ô)
    ws.append([])
    row = ws.max_row + 1
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "TOA DO COC"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align

    # Hàng tiêu đề cột và danh sách tọa độ từng cọc
    ws.append(["Coc", "X (m)", "Y (m)"])
    for cell in ws[ws.max_row]: cell.font = bold_font

    for i, (x, y) in enumerate(config['coords']):
        ws.append([i+1, round(x,3), round(y,3)])

    wb.save(out_path)
    return out_path


# ============================================================================
# XUẤT PDF
# ============================================================================
def export_pdf(config, loads, params, out_dir, prefix, png_path=None):
    """
    Xuất báo cáo tóm tắt phương án cọc ra tệp PDF (1 trang).

    Tham số:
        config          : phương án kiến nghị (n, type, pmax, pmin...).
        loads           : danh sách tổ hợp tải trọng.
        params          : thông số đầu vào (kích thước bệ, [Po], [Ct]...).
        out_dir, prefix : thư mục đích và tiền tố tên tệp.
        png_path        : (tùy chọn) đường dẫn ảnh mặt bằng để nhúng vào PDF.
    Tạo tệp '<prefix>_report.pdf' gồm thông số đầu vào, phương án kiến nghị
    và (nếu có) hình mặt bằng. Trả về đường dẫn tệp đã ghi.
    """
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    out_path = os.path.join(out_dir, f"{prefix}_report.pdf")
    c = pdf_canvas.Canvas(out_path, pagesize=A4)
    width, height = A4

    # Tiêu đề báo cáo (canh giữa, đầu trang)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width/2.0, height - 50, f"BAO CAO TOI UU HOA - {prefix}")

    c.setFont("Helvetica", 12)
    y = height - 80

    # 1. Thông số đầu vào
    c.drawString(50, y, "1. Thong so dau vao:"); y -= 20
    c.drawString(70, y, f"Kich thuoc be: {params.get('L_X')} x {params.get('L_Y')} m"); y -= 15
    c.drawString(70, y, f"Duong kinh coc: {params.get('D_PILE')} m"); y -= 15
    c.drawString(70, y, f"Suc chiu tai cho phep: Po = {params.get('P_LIMIT')} T, Ct = {params.get('P_TENSION')} T"); y -= 15
    if params.get('M_LIMIT', 0) > 0:
        c.drawString(70, y, f"Suc uon cho phep: Mmax = {params.get('M_LIMIT')} T.m"); y -= 15
    else:
        y -= 15

    # 2. Phương án kiến nghị
    c.drawString(50, y, "2. Phuong an kien nghi:"); y -= 20
    c.drawString(70, y, f"So luong coc: {config.get('n', 0)} coc"); y -= 15
    config_type = config.get('type', '')
    if config_type == 'Gốc': config_type = 'Goc'
    c.drawString(70, y, f"Kieu bo tri: {config_type}"); y -= 15
    c.drawString(70, y, f"Pmax = {config.get('pmax', 0):.2f} T"); y -= 15
    c.drawString(70, y, f"Pmin = {config.get('pmin', 0):.2f} T"); y -= 15
    if params.get('M_LIMIT', 0) > 0:
        c.drawString(70, y, f"Mmax = {max(config.get('mxmax', 0), config.get('mymax', 0)):.2f} T.m"); y -= 15
    c.drawString(70, y, f"Trang thai: {'DAT' if config.get('ok') else 'KHONG DAT'}"); y -= 15
    c.drawString(70, y, f"Ly do: {config.get('msg', '')}"); y -= 30

    # 3. Hình mặt bằng (chỉ nhúng khi có ảnh PNG hợp lệ)
    if png_path and os.path.exists(png_path):
        c.drawString(50, y, "3. Hinh anh mat bang:"); y -= 10
        # Nhúng ảnh, co theo bề rộng nhưng giữ tỉ lệ khung hình
        from reportlab.lib.utils import ImageReader
        img = ImageReader(png_path)
        orig_w, orig_h = img.getSize()
        aspect = orig_h / float(orig_w)
        img_w = 450
        img_h = img_w * aspect

        # Bảo đảm ảnh không tràn xuống đáy trang
        if img_h > y - 50:
            img_h = y - 50
            img_w = img_h / aspect

        c.drawImage(png_path, (width - img_w)/2.0, y - img_h, width=img_w, height=img_h)
        y -= (img_h + 30)

    c.showPage()
    c.save()
    return out_path
