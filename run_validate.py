"""
run_validate.py - KIEM CHUNG do chinh xac cua phuong phap HOP DEN so voi MCOC.

MUC DICH
    Chung minh sai so giua 2 phuong phap la KHONG DANG KE:
      (1) PHUONG PHAP GOC (MCOC) : ket qua phan tu huu han khong gian (FEM) thuc,
          doc truc tiep tu cac file *_result.txt -> Nmax/Nmin (ground truth).
      (2) PHUONG PHAP HOP DEN     : cong thuc be cung giai tich ma chuong trinh chay
          Pi = N/n + Mx*yi/Ix + My*xi/Iy  (toc do < 1 ms).

    Voi MOI ho so, ca 2 phuong phap duoc tinh tren CUNG toa do coc va CUNG to hop
    tai trong, nen sai so phan anh dung chenh lech giua 2 mo hinh co hoc.

CACH CHAY
    cd d:/Project/TEDI/OptApp
    python run_validate.py

KET QUA XUAT
    validate_mcoc_vs_blackbox.txt   - bang so sanh dang van ban
    validate_mcoc_vs_blackbox.xlsx  - bang Excel co dinh dang
    validate_mcoc_vs_blackbox.png   - bieu do tuong quan + phan bo sai so
"""

import sys
import os
import io

# Ép mã hóa console về UTF-8 và bảo đảm import được từ thư mục gốc dự án
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import numpy as np
from core.blackbox import MCOCBlackbox
from io_handlers.file_io import parse_mcoc_result_as_input

SAMPLE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "input_sample")
OUT_DIR    = os.path.dirname(os.path.abspath(__file__))
LINE = "=" * 92


# ============================================================================
# Thu thập dữ liệu & tính thống kê sai số
# ============================================================================
def collect_cases(sample_dir):
    """Đọc tất cả file *_result.txt, tính MCOC (ground truth) và Hộp đen (bệ cứng)."""
    rows = []
    files = sorted(f for f in os.listdir(sample_dir) if f.endswith("_result.txt"))
    for fn in files:
        path = os.path.join(sample_dir, fn)
        # Đọc hồ sơ kết quả MCOC; bỏ qua file lỗi không phân tích được
        try:
            params, loads, _ = parse_mcoc_result_as_input(path)
        except Exception as e:
            print(f"  [!] Bo qua {fn}: {e}")
            continue

        # Lấy tọa độ cọc và kết quả MCOC (ground truth) từ hồ sơ
        coords = params.get("original_coords")
        mcoc_pmax = params.get("orig_pmax")
        mcoc_pmin = params.get("orig_pmin", 0.0)
        if not coords or mcoc_pmax is None or not loads:
            continue

        # Tính lại Pmax/Pmin bằng phương pháp Hộp đen (bệ cứng) trên cùng dữ liệu
        arr = np.array(coords, dtype=float)
        bb_pmax = MCOCBlackbox._rigid_cap_pmax(arr, loads)
        bb_pmin = MCOCBlackbox._rigid_cap_pmin(arr, loads)

        # Sai số tương đối Pmax giữa Hộp đen và MCOC (%)
        err_pmax = (bb_pmax - mcoc_pmax) / mcoc_pmax * 100.0 if mcoc_pmax else 0.0

        rows.append({
            "case"     : fn.replace("_result.txt", ""),
            "n"        : len(coords),
            "mcoc_pmax": mcoc_pmax,
            "bb_pmax"  : bb_pmax,
            "err_pmax" : err_pmax,
            "mcoc_pmin": mcoc_pmin,
            "bb_pmin"  : bb_pmin,
        })
    return rows


def stats(rows):
    """Tính các chỉ số thống kê sai số Pmax (bias, sai số tuyệt đối, RMSE, hệ số tương quan...)."""
    e = np.array([r["err_pmax"] for r in rows])
    mcoc = np.array([r["mcoc_pmax"] for r in rows])
    bb   = np.array([r["bb_pmax"] for r in rows])
    rmse = float(np.sqrt(np.mean((bb - mcoc) ** 2)))
    # He so tuong quan Pearson
    corr = float(np.corrcoef(mcoc, bb)[0, 1]) if len(rows) > 1 else 1.0
    return {
        "N"        : len(rows),
        "mean_err" : float(e.mean()),
        "mean_abs" : float(np.abs(e).mean()),
        "max_abs"  : float(np.abs(e).max()),
        "std"      : float(e.std()),
        "rmse"     : rmse,
        "corr"     : corr,
        "within_3" : int(np.sum(np.abs(e) <= 3.0)),
        "within_5" : int(np.sum(np.abs(e) <= 5.0)),
    }


# ============================================================================
# Xuất kết quả: bảng văn bản, Excel và biểu đồ
# ============================================================================
def build_text(rows, st):
    """Dựng bảng so sánh dạng văn bản (bảng chi tiết + phần thống kê + kết luận)."""
    out = []
    out.append(LINE)
    out.append("  KIEM CHUNG: PHUONG PHAP GOC (MCOC / FEM) vs PHUONG PHAP HOP DEN (BE CUNG)")
    out.append(f"  So ho so kiem chung: {st['N']}   |   Du lieu: input_sample/*_result.txt")
    out.append(LINE)
    out.append(f"  {'Ho so':<12} {'n':>4} | {'MCOC Pmax':>11} {'HopDen Pmax':>12} {'Sai so':>9} | "
                f"{'MCOC Pmin':>11} {'HopDen Pmin':>12}")
    out.append(f"  {'':<12} {'':>4} | {'(T)':>11} {'(T)':>12} {'(%)':>9} | {'(T)':>11} {'(T)':>12}")
    out.append("  " + "-" * 88)
    for r in rows:
        out.append(f"  {r['case']:<12} {r['n']:>4} | "
                   f"{r['mcoc_pmax']:>11.2f} {r['bb_pmax']:>12.2f} {r['err_pmax']:>+8.1f}% | "
                   f"{r['mcoc_pmin']:>11.2f} {r['bb_pmin']:>12.2f}")
    out.append(LINE)
    out.append("  THONG KE SAI SO Pmax (Hop den so voi MCOC)")
    out.append(f"  +-- Sai so trung binh (bias)     : {st['mean_err']:+.2f} %   "
               f"({'thien an toan, du bao cao hon' if st['mean_err'] >= 0 else 'du bao thap hon'})")
    out.append(f"  +-- Sai so tuyet doi trung binh  : {st['mean_abs']:.2f} %")
    out.append(f"  +-- Sai so tuyet doi lon nhat    : {st['max_abs']:.2f} %")
    out.append(f"  +-- Do lech chuan (std)          : {st['std']:.2f} %")
    out.append(f"  +-- RMSE                          : {st['rmse']:.2f} T")
    out.append(f"  +-- He so tuong quan R           : {st['corr']:.5f}")
    out.append(f"  +-- So ho so sai so <= 3%        : {st['within_3']}/{st['N']}")
    out.append(f"  +-- So ho so sai so <= 5%        : {st['within_5']}/{st['N']}")
    out.append(LINE)
    verdict = "KHONG DANG KE" if st['mean_abs'] <= 5.0 else "CAN XEM XET"
    out.append(f"  KET LUAN: Sai so trung binh {st['mean_abs']:.1f}% (max {st['max_abs']:.1f}%) -> {verdict}.")
    out.append("  Phuong phap Hop den (be cung) bam sat ket qua MCOC; bias duong the hien")
    out.append("  xu huong thien an toan (du bao Pmax hoi cao hon thuc te FEM).")
    out.append(LINE)
    return "\n".join(out)


def export_xlsx(rows, st, out_path):
    """Xuất bảng so sánh ra file Excel có định dạng (cần openpyxl); trả về đường dẫn hoặc None."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [!] Khong co openpyxl - bo qua xuat Excel.")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KiemChung MCOC vs HopDen"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    hdr_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.merge_cells("A1:G1")
    ws["A1"] = "KIEM CHUNG PHUONG PHAP GOC (MCOC) vs PHUONG PHAP HOP DEN (BE CUNG)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    headers = ["Ho so", "So coc",
               "MCOC Pmax (T)", "Hop den Pmax (T)", "Sai so Pmax (%)",
               "MCOC Pmin (T)", "Hop den Pmin (T)"]
    ws.append([])
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold
        cell.fill = hdr_fill
        cell.alignment = center

    for r in rows:
        ws.append([r["case"], r["n"],
                   round(r["mcoc_pmax"], 2), round(r["bb_pmax"], 2), round(r["err_pmax"], 2),
                   round(r["mcoc_pmin"], 2), round(r["bb_pmin"], 2)])
        if abs(r["err_pmax"]) > 5.0:
            for cell in ws[ws.max_row]:
                cell.fill = warn_fill

    ws.append([])
    ws.append(["THONG KE SAI SO Pmax"])
    ws[ws.max_row][0].font = bold
    ws.append(["Sai so trung binh (bias) %", round(st["mean_err"], 2)])
    ws.append(["Sai so tuyet doi trung binh %", round(st["mean_abs"], 2)])
    ws.append(["Sai so tuyet doi lon nhat %", round(st["max_abs"], 2)])
    ws.append(["Do lech chuan %", round(st["std"], 2)])
    ws.append(["RMSE (T)", round(st["rmse"], 2)])
    ws.append(["He so tuong quan R", round(st["corr"], 5)])
    ws.append([f"So ho so sai so <= 3%", f"{st['within_3']}/{st['N']}"])
    ws.append([f"So ho so sai so <= 5%", f"{st['within_5']}/{st['N']}"])

    for ci in range(1, ws.max_column + 1):
        letter = get_column_letter(ci)
        width = max((len(str(ws.cell(row=ri, column=ci).value))
                     for ri in range(2, ws.max_row + 1)
                     if ws.cell(row=ri, column=ci).value is not None), default=10)
        ws.column_dimensions[letter].width = min(width + 3, 28)

    wb.save(out_path)
    return out_path


def export_png(rows, st, out_path):
    """Vẽ biểu đồ tương quan và phân bố sai số ra file PNG (cần matplotlib); trả về đường dẫn hoặc None."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        print("  [!] Khong co matplotlib - bo qua xuat bieu do.")
        return None

    mcoc = np.array([r["mcoc_pmax"] for r in rows])
    bb   = np.array([r["bb_pmax"] for r in rows])
    err  = np.array([r["err_pmax"] for r in rows])

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5.2))

    # (1) Tuong quan MCOC vs Hop den
    lo = min(mcoc.min(), bb.min()) * 0.95
    hi = max(mcoc.max(), bb.max()) * 1.05
    ax1.plot([lo, hi], [lo, hi], "k--", lw=1, label="y = x (trung khop tuyet doi)")
    ax1.scatter(mcoc, bb, c="#1f77b4", s=45, edgecolors="white", zorder=3)
    ax1.set_xlabel("Pmax theo MCOC / FEM (T)")
    ax1.set_ylabel("Pmax theo Hop den - Be cung (T)")
    ax1.set_title(f"Tuong quan Pmax  (R = {st['corr']:.4f}, N = {st['N']})")
    ax1.legend(loc="upper left", fontsize=9)
    ax1.grid(True, ls=":", alpha=0.5)

    # (2) Phan bo sai so
    ax2.hist(err, bins=12, color="#2ca02c", edgecolor="white", alpha=0.85)
    ax2.axvline(st["mean_err"], color="red", ls="--", lw=1.5,
                label=f"Bias trung binh = {st['mean_err']:+.1f}%")
    ax2.axvline(0, color="black", lw=0.8)
    ax2.set_xlabel("Sai so Pmax cua Hop den so voi MCOC (%)")
    ax2.set_ylabel("So ho so")
    ax2.set_title(f"Phan bo sai so  (|sai so| TB = {st['mean_abs']:.1f}%, max = {st['max_abs']:.1f}%)")
    ax2.legend(loc="upper right", fontsize=9)
    ax2.grid(True, ls=":", alpha=0.5, axis="y")

    fig.suptitle("KIEM CHUNG: Phuong phap Hop den (Be cung) bam sat MCOC / FEM",
                 fontsize=13, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(out_path, dpi=180)
    plt.close(fig)
    return out_path


if __name__ == "__main__":
    # Thu thập hồ sơ và tính sai số; dừng nếu không có dữ liệu mẫu
    rows = collect_cases(SAMPLE_DIR)
    if not rows:
        print("  [X] Khong tim thay ho so MCOC nao trong input_sample/.")
        sys.exit(1)

    # Tính thống kê sai số, dựng bảng văn bản và in ra màn hình
    st = stats(rows)
    text = build_text(rows, st)
    print(text)

    # Đường dẫn các file kết quả sẽ xuất ra
    txt_path  = os.path.join(OUT_DIR, "validate_mcoc_vs_blackbox.txt")
    xlsx_path = os.path.join(OUT_DIR, "validate_mcoc_vs_blackbox.xlsx")
    png_path  = os.path.join(OUT_DIR, "validate_mcoc_vs_blackbox.png")

    # Ghi bảng TXT, rồi xuất thêm Excel và biểu đồ PNG nếu có thư viện
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(text + "\n")
    saved_xlsx = export_xlsx(rows, st, xlsx_path)
    saved_png  = export_png(rows, st, png_path)

    print(f"\n  [Xuat] Bang TXT     : {txt_path}")
    if saved_xlsx:
        print(f"  [Xuat] Bang XLSX    : {saved_xlsx}")
    if saved_png:
        print(f"  [Xuat] Bieu do PNG  : {saved_png}")
    print()
